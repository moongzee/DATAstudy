import pandas as pd
import numpy as np
import glob
from openpyxl import load_workbook

## 코드 용어 시스템명 check function
def code_test(unuse_yn, postfix, postfix_chk, ko_1, DB_name):
    if unuse_yn =='':
        if (postfix == 'CD') & (postfix_chk =='CD'):
            if str(ko_1) != str(DB_name):
                return 'O'
            else:
                return 'X'
        else :
            return 'X'
    else:
        return 'X'

## 분류어가 없는 경우 check function
def not_in_postfix(unuse_yn, standard_ko_term, postfix_chk):
    if unuse_yn =='':
        if (standard_ko_term != '') & (postfix_chk == '') :
            return 'O'
        else :
            return 'X'
    else:
        return 'X'

## 한단어만 나온 경우 check function
def extract_one_word(unuse_yn, ko_1, ko_2, ko_3, ko_4, ko_5, ko_6):
    if unuse_yn == '':
        if (ko_1 != '') & (ko_2=='') & (ko_3=='') & (ko_4=='') & (ko_5=='') & (ko_6 ==''):
            return 'O'
        else:     
            return 'X'
    else:
        return 'X'


## flag yn check function 
def flag_yn(unuse_yn, en_col, postfix_chk):
    if unuse_yn == '':
        if ('FLAG' in en_col) & (postfix_chk !='YN'):
            return 'O'
        else :
            return 'X'
    else:
        return 'X'


## eai flag yn check function 
def EAI_flag_yn(unuse_yn, en_col, standard_ko_term):
    if unuse_yn == '':
        if (en_col == 'EAI_FLAG') & (standard_ko_term !='EAI연계여부'):
            return 'O'
        else :
            return 'X'
    else:
        return 'X'

    
## DML check function
def DML_code_chk(unuse_yn, en_col, standard_ko_term):
    if unuse_yn =='':
        if ('DML' in en_col) & (standard_ko_term !='EAIDML코드'):
            return 'O'
        else :
            return 'X'
    else:
        return 'X'
    
    
## duplicated_double_chk function 
def df_dup_double_chk( unuse_yn, tmp_col, target_col , filter_col_list):
    if unuse_yn == '':
        if tmp_col == 'X' :
            if  target_col in filter_col_list : 
                    return 'O'
            else : 
                return 'X'
        else : 
            return 'O'
    else:
        return 'X'    

  
  
  
if __name__ == "__main__":
    
    file_list = glob.glob('asset/*.xlsx')

    print(f'test file list : {file_list}')

    for i in file_list:
        file_path = i.replace("\\","/")

        #open excel sheet 
        wb = load_workbook(filename=file_path, data_only=True)
        sheetname = wb.sheetnames[1]
        
        df = pd.read_excel(file_path, sheet_name=sheetname)
        print(f'read_excel_file : {file_path} | Sheet name : {sheetname}')
        tmp_df = df.rename(columns=df.iloc[0])
        extract_df = tmp_df.drop(tmp_df.index[0]) 
        extract_df.fillna('', inplace=True)


        # check list data frame create    
        print('code_correct_check')
        extract_df['시스템코드체크'] = extract_df.apply(lambda row : code_test(row['미사용구분'],row['분류어'],row['분류어chk'],row['한글1'],row['DB명']), axis=1)

        print('postfix check')
        extract_df['분류어X'] = extract_df.apply(lambda row : not_in_postfix(row['미사용구분'],row['한글용어명'],row['분류어chk']), axis=1)

        print('extract one word')
        extract_df['한단어'] = extract_df.apply(lambda row : extract_one_word(row['미사용구분'],row['한글1'],row['한글2'],row['한글3'],row['한글4'],row['한글5'],row['한글6']), axis=1)

        print('check flag_yn')
        extract_df['FLAG_여부X'] = extract_df.apply(lambda row : flag_yn(row['미사용구분'],row['컬럼명(영문)'], row['분류어chk']), axis=1)

        print('check EAI_flag_yn')
        extract_df['EAI_FLAG_여부_체크'] = extract_df.apply(lambda row : EAI_flag_yn(row['미사용구분'],row['컬럼명(영문)'], row['한글용어명']), axis=1)

        #print('check DML code')
        #extract_df['DML_code_chk'] = extract_df.apply(lambda row : DML_code_chk(row['컬럼명(영문)'], row['한글용어명']), axis=1)

        print('check duplicated columns in 1 table')
        df_dup_chk1 = extract_df.duplicated(['DB명','테이블명(영문)','한글용어명','미사용구분'],keep=False)&(extract_df['한글용어명']!='')&(extract_df['미사용구분']=='')
        extract_df['테이블컬럼중복'] = np.where((df_dup_chk1), 'O', 'X')

        print('check different ko column in 1 en column')    
        df_dup_chk2 =(~(extract_df.duplicated(['컬럼명(영문)','한글용어명'],keep=False)))&(extract_df['한글용어명']!='')&(extract_df['미사용구분']=='')&(extract_df.duplicated(['컬럼명(영문)'],keep=False))
        extract_df['tmp_col'] = np.where((df_dup_chk2), 'O', 'X')
        filter_col_list = extract_df[extract_df['tmp_col']=='O']['컬럼명(영문)'].to_list()    
        extract_df['영문컬럼기준_한글용어다름'] = extract_df.apply(lambda row : df_dup_double_chk( row['미사용구분'], row['tmp_col'], row['컬럼명(영문)'], filter_col_list) , axis=1)
        
        extract_df.drop(columns=['tmp_col'], axis=1, inplace=True)
        filter_col_list.clear()

        print('check different en column in 1 ko column')
        df_dup_chk3 = (~(extract_df.duplicated(['한글용어명','영문용어명'],keep=False)))&(extract_df['한글용어명']!='')&(extract_df['미사용구분']=='')&(extract_df.duplicated(['한글용어명'],keep=False))
        extract_df['tmp_col'] = np.where((df_dup_chk3), 'O', 'X') 
        filter_col_list = extract_df[extract_df['tmp_col']=='O']['한글용어명'].to_list()
        extract_df['한글용어기준_영문용어다름'] = extract_df.apply(lambda row : df_dup_double_chk( row['미사용구분'], row['tmp_col'], row['한글용어명'], filter_col_list) , axis=1)
        
        extract_df.drop(columns=['tmp_col'], axis=1, inplace=True)
        filter_col_list.clear()


        print('****** result check df ******')
        checklist_df = extract_df[['시스템코드체크','분류어X','한단어','FLAG_여부X','EAI_FLAG_여부_체크','테이블컬럼중복','영문컬럼기준_한글용어다름','한글용어기준_영문용어다름']]



        print(checklist_df)

        print('Excel UPDATE progress start')
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')


        checklist_df.to_excel(writer,
                              sheet_name=sheetname,
                              index=False, 
                              startrow=1,
                              startcol=38)


        writer.close()
        print(f'Excel UPDATE progress end. The filepath: {file_path}')