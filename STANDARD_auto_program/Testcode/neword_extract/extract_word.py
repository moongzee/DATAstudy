import pandas as pd 
import numpy as np
import glob
from openpyxl import load_workbook



  
if __name__ == "__main__":
    
    file_list = glob.glob('asset/*.xlsx')
    
    print(f'test file list : {file_list}')

    
    for i in file_list : 
        file_path = i.replace("\\","/")
    
        # open excel sheet
        wb = load_workbook(filename=file_path, data_only=True)
        sheetname = wb.sheetnames[1]
        
        df = pd.read_excel(file_path, sheet_name=sheetname)
        
        new_word_df = df[['단어추가대상-자동생성', 'Unnamed: 26','Unnamed: 27', 'Unnamed: 28', 'Unnamed: 29', 'Unnamed: 30']]
        df.drop(df.index, inplace=True)
        new_word_df = new_word_df.drop(0)
        
        
        tmp_df = pd.concat([new_word_df['단어추가대상-자동생성'].dropna(axis=0), 
                            new_word_df['Unnamed: 26'].dropna(axis=0),
                            new_word_df['Unnamed: 27'].dropna(axis=0),
                            new_word_df['Unnamed: 28'].dropna(axis=0),
                            new_word_df['Unnamed: 29'].dropna(axis=0),
                            new_word_df['Unnamed: 30'].dropna(axis=0)       ])
        
        tmp_list = []
        
        for i in tmp_df :
            tmp_list.append(i)
            
            
        result = sorted(list(set(tmp_list)))
        
        result_df = pd.DataFrame()
        result_df['new_word']=result[1:]

        
        file_name = file_path.replace('.xlsx','_new_word_list.csv')
        result_file_name = file_name.replace('asset/','result/')
        result_df.to_csv(result_file_name, encoding = 'cp949')