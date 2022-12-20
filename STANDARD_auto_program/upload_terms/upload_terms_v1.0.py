import pandas as pd
import shutil
import glob
import math
from openpyxl import load_workbook


## LONG, NVARCHAR2, TIMESTAMP(6), CLOB, BLOB -> VARCHAR2
def convert_datatype (datatype):
    convert_list = ['LONG','NVARCHAR2','TIMESTAMP(6)','CLOB','BLOB']
        
    if datatype in convert_list :
        return 'VARCHAR2'
    else :
        return datatype


## 수정될 수 있음---> 고치자
def find_nearlist_val (data_length_list, data_length):
    if data_length in [None, ''] :
        return ''
    else:
        new_list = [x for x in data_length_list ]
        if len(new_list) == 0 :
                return '' 
        else : 
            data_length = float(data_length)
            minValue = min(new_list, key = lambda x : abs(x-data_length))
            return minValue
        


def extract_cd_domain (cd_domain_df , target_en_term ):
    word_list = target_en_term.split('_')
    for i in range(len(word_list)):
        comp_str = '_'.join(word_list[i:])
        if  comp_str in cd_domain_df['도메인영문명'].to_list():
            select_domain = cd_domain_df[cd_domain_df['도메인영문명']==comp_str]['도메인명'].to_list()[0]  
            break
        else : 
            select_domain = ''
    return select_domain 



def select_domin (domain_df, cd_domain_df, target_last_word, target_en_term, data_type, data_length):
    # 코드도메인 불러오는 로직 짜야함
    if target_last_word in domain_df['분류어'].to_list() : 
        if target_last_word == 'CD':
            return  extract_cd_domain(cd_domain_df, target_en_term)
        else:
            data_type = convert_datatype(data_type)    
            tmp_df = domain_df[domain_df['분류어']==target_last_word]
            if len(tmp_df['데이터타입'].to_list()) != 1 :
                select_data_length = find_nearlist_val(sorted(tmp_df[tmp_df['데이터타입']==data_type]['데이터길이'].to_list()), data_length)
                if  select_data_length == '' or math.isnan(float(select_data_length)):     
                    return ''
                else : 
                    domain = tmp_df[(tmp_df['데이터타입']==data_type)&(tmp_df['데이터길이']==select_data_length)]['도메인명'].to_list()[0]
                    return domain
            else :         
                domain = tmp_df['도메인명'].to_list()[0]
                return domain                  
    else : 
        return ''


    
    
def make_meta_df (file_path):
    wb = load_workbook(filename=file_path, data_only=True)
    sheetname = wb.sheetnames[1]

    df = pd.read_excel(file_path, sheet_name=sheetname)
    print(f'read_excel_file : {file_path} | Sheet name : {sheetname}')
    
    tmp_df = df.rename(columns=df.iloc[0])
    extract_df = tmp_df.drop(tmp_df.index[0]) 
    extract_df.fillna('', inplace=True)

    #create meta data frame
    meta_df = extract_df[(extract_df['미등록용어']!='')&(extract_df['분류어chk']!='CD')&(extract_df['미사용구분']=='')&(extract_df['분류어chk']!='')]        
    meta_df = meta_df[['한글용어명','영문용어명','분류어chk','데이터 타입','데이터 길이']]

    meta_df = meta_df.astype({  '한글용어명' : 'string',
                                '영문용어명' : 'string',
                                '분류어chk' : 'string',
                                '데이터 타입' : 'string',
                                '데이터 길이' : 'string'    })
    
    meta_df['데이터 길이'] = meta_df['데이터 길이'].apply(lambda x : x.replace(',','.'))
    meta_df = meta_df.astype({ '데이터 길이' : 'float' })
    
    return meta_df
    
    
    
### 로직 추가해야함     
def select_col_length (select_length_df, target_last_word, data_type) :
    if target_last_word in select_length_df['분류어chk'].to_list() :
        data_type = convert_datatype(data_type)
        if data_type in select_length_df[select_length_df['분류어chk']==target_last_word]['데이터 타입'].to_list():
            filter_data_df = select_length_df[(select_length_df['분류어chk']==target_last_word)&(select_length_df['데이터 타입']==data_type)]
            max_cnt = filter_data_df['빈도수'].max()
            chk_df = filter_data_df[filter_data_df['빈도수']==max_cnt]
            if len(chk_df) !=1 :
                data_length = chk_df['데이터 길이'].to_list()[0]
            else :
                data_length = chk_df['데이터 길이'].max()
        else :
            data_length= ''
    else :
        data_length= ''
    return data_length


def select_cd_data_type (cd_domain_df, domain_name) :
    if domain_name in cd_domain_df['도메인명'].to_list() :
        return cd_domain_df[cd_domain_df['도메인명']==domain_name]['논리데이터타입'].to_list()[0]
    else : 
        pass



if __name__ == "__main__":
    
    file_list = glob.glob('asset/meta/표준화 작업*.xlsx')
    dt_list = glob.glob('asset/data/*.xlsx')
    print(f'upload target meta file list : {file_list}')

    ## 기준 도메인 파일 불러오기
    domain_file_path = 'C:/Users/moonj/Desktop/남동발전표준화/자동화/upload_terms/asset/std/term_domain.xlsx'
    cd_domain_file_path = 'C:/Users/moonj/Desktop/남동발전표준화/자동화/upload_terms/asset/std/cd_domain.xlsx'
    print(f'read term_domain file: {domain_file_path}')
    print(f'read cd_domain file: {cd_domain_file_path}')
    
    domain_df = pd.read_excel(domain_file_path, sheet_name='분류어&도메인 매핑')
    cd_domain_df = pd.read_excel(cd_domain_file_path, sheet_name='도메인목록')
    
    domain_df = domain_df.astype({  '분류어':'string',
                                    '단어명':'string',
                                    '도메인명':'string',
                                    '데이터타입':'string',
                                    '데이터길이':'string'   })
    domain_df.fillna('',inplace=True)
    domain_df['데이터길이']=domain_df['데이터길이'].apply(lambda x : x.replace(',','.'))
    domain_df['데이터길이']=domain_df.apply(lambda row : None if row['데이터길이']=='' else row['데이터길이'], axis=1)
    domain_df = domain_df.astype({'데이터길이':'float'})


    cd_domain_df = cd_domain_df[['도메인명','도메인영문명','논리데이터타입']]
    cd_domain_df = cd_domain_df.astype({'도메인명' : 'string', '도메인영문명':'string', '논리데이터타입':'string'})
    cd_domain_df.fillna('',inplace=True)



    for i,j in enumerate(zip(file_list, dt_list)):
        file_path = j[0].replace("\\","/")
        dt_path = j[1].replace("\\", "/")
        
        
        #create meta data frame
        meta_df = make_meta_df(file_path)
        
        #create target_data frame
        target_df = pd.read_excel(dt_path)
        target_df.fillna('',inplace=True)
        
        
        #### 컬럼길이 가져오기
        ## 다시 할것           
        select_length_df = meta_df.groupby(['분류어chk','데이터 타입','데이터 길이'])['한글용어명'].count().reset_index()
        select_length_df.rename(columns={'한글용어명':'빈도수'}, inplace=True)



        target_df['데이터 길이']=target_df.apply(lambda row : select_col_length(select_length_df, row['분류어chk'], row['데이터 타입']) if row['데이터 길이']=='' else row['데이터 길이'], axis=1)
        target_df=target_df.astype({'데이터 길이':'string'})
        target_df['데이터 길이']=target_df['데이터 길이'].apply(lambda x : x.replace(',','.'))
        target_df['데이터 길이']=target_df.apply(lambda row : None if row['데이터 길이']=='' else row['데이터 길이'], axis=1)
        #target_df = target_df.astype({'데이터 길이':'float'})
        target_df['도메인명']=target_df.apply(lambda row : select_domin(domain_df, cd_domain_df, row['분류어chk'], row['영문용어명'], row['데이터 타입'], row['데이터 길이']), axis=1)
        target_df['데이터 타입']=target_df.apply(lambda row : select_cd_data_type(cd_domain_df, row['도메인명']) if row['데이터 타입']=='' else row['데이터 타입'], axis=1)

        
        
        #create upload data frame
        upload_df = pd.DataFrame(columns=['요청유형','용어명','용어영문명','도메인명','표준여부','용어설명','*데이터타입','*데이터길이'])
        upload_df['용어명']=target_df['한글용어명']
        upload_df['용어영문명']=target_df['영문용어명']
        upload_df['도메인명']=target_df['도메인명']
        upload_df['용어설명']=target_df['한글용어명']
        upload_df['*데이터길이']=target_df['데이터 길이']
        upload_df['*데이터타입']=target_df['데이터 타입']
        upload_df['요청유형']='신규'
        upload_df['표준여부']=upload_df.apply(lambda x : 'Y' if x['도메인명']!='' else '', axis=1)
        
        # clear dataframe
        
        target_df = target_df[0:0]
        select_length_df = select_length_df[0:0]
        meta_df = meta_df[0:0]
        

        #create upload file
        upload_file_sc = glob.glob('asset/sample/*.xlsx')
        upload_file = upload_file_sc[0].replace('\\','/')
        upload_file = upload_file.replace('sample','result')
        upload_file = upload_file.replace('용어등록 샘플.xlsx',file_path.split('/')[-1])
        
        shutil.copy(upload_file_sc[0],upload_file)

        print('Excel UPDATE progress start')
        writer = pd.ExcelWriter(upload_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')

        upload_df.to_excel  (   writer,
                                sheet_name='용어목록',
                                header=False,
                                index=False, 
                                startrow=1,
                                startcol=1      )

        writer.close()

